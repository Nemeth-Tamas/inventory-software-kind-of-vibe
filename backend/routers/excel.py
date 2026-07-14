from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from typing import Optional
from database import get_db
from models import Product
from models_inventory import Stocktake, StocktakeItem, InventoryMovement
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import io

router = APIRouter(prefix="/api/excel", tags=["excel"])


@router.get("/export/products")
async def export_products(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Product).where(Product.is_archived.is_(False)))
    products = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Termékek"

    headers = [
        "Belső vonalkód",
        "EAN",
        "Termék név",
        "Cikkszám (SKU)",
        "Gyártói SKU",
        "Egység",
        "ÁFA (%)",
        "Nettó beszerzési ár (Ft)",
        "Bruttó eladási ár (Ft)",
        "Készlet",
    ]
    ws.append(headers)

    # Style header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="1E293B", end_color="1E293B", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")

    for p in products:
        ws.append(
            [
                p.barcode,
                p.ean or "",
                p.name,
                p.sku,
                p.manufacturer_sku or "",
                p.unit,
                p.vat_rate,
                p.purchase_price_net,
                p.sale_price_gross,
                p.current_stock,
            ]
        )

    # Set column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Freeze header
    ws.freeze_panes = "A2"

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=termekek.xlsx"},
    )


@router.get("/export/stocktake/{stocktake_id}")
async def export_stocktake(stocktake_id: str, db: AsyncSession = Depends(get_db)):
    # Query stocktake and items
    st_res = await db.execute(select(Stocktake).where(Stocktake.id == stocktake_id))
    st = st_res.scalar_one_or_none()
    if not st:
        raise HTTPException(status_code=404, detail="Leltár nem található")

    items_res = await db.execute(
        select(StocktakeItem).where(StocktakeItem.stocktake_id == stocktake_id)
    )
    items = items_res.scalars().all()

    wb = openpyxl.Workbook()

    ws_all = wb.active
    ws_all.title = "Teljes leltár"
    headers = [
        "Vonalkód",
        "Termék név",
        "Mennyiség",
        "Egységár nettó (Ft)",
        "Összérték nettó (Ft)",
    ]
    ws_all.append(headers)

    for item in items:
        p_res = await db.execute(select(Product).where(Product.id == item.product_id))
        p = p_res.scalar_one()
        ws_all.append(
            [
                p.barcode,
                p.name,
                item.counted_qty,
                p.purchase_price_net,
                item.counted_qty * p.purchase_price_net,
            ]
        )

    # Style Teljes leltár header
    for col_num, header in enumerate(headers, 1):
        cell = ws_all.cell(row=1, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="1E293B", end_color="1E293B", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")

    # Set column widths
    for col in ws_all.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_all.column_dimensions[col_letter].width = max(max_len + 3, 12)

    ws_all.freeze_panes = "A2"

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=leltar_{stocktake_id}.xlsx"
        },
    )


@router.get("/export/valuation")
async def export_valuation(
    category_id: Optional[str] = None,
    location_id: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
):
    stmt = select(Product).where(Product.is_archived.is_(False))
    if category_id:
        stmt = stmt.where(Product.category_id == category_id)
    if location_id:
        stmt = stmt.where(Product.default_location_id == location_id)
        
    stmt = stmt.options(
        selectinload(Product.category),
        selectinload(Product.default_location)
    )
    
    result = await db.execute(stmt)
    products = result.scalars().all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Készletérték"
    
    headers = [
        "Terméknév",
        "SKU",
        "Kategória",
        "Helyszín",
        "Készlet (db)",
        "Nettó beszerzési egységár (Ft)",
        "Bruttó beszerzési egységár (Ft)",
        "Összesített nettó érték (Ft)",
        "Összesített bruttó érték (Ft)"
    ]
    ws.append(headers)
    
    # Style header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="1E293B", end_color="1E293B", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")
        
    total_stock = 0
    total_net = 0
    total_gross = 0
    
    for p in products:
        stock = p.current_stock or 0
        price_net = p.purchase_price_net or 0
        price_gross = p.purchase_price_gross or 0
        
        if price_net > 0 and price_gross == 0:
            price_gross = int(price_net * 1.27)
        elif price_gross > 0 and price_net == 0:
            price_net = int(price_gross / 1.27)
            
        tot_net = stock * price_net
        tot_gross = stock * price_gross
        
        ws.append([
            p.name,
            p.sku,
            p.category.name if p.category else "Nincs kategória",
            p.default_location.name if p.default_location else "Nincs helyszín",
            stock,
            price_net,
            price_gross,
            tot_net,
            tot_gross
        ])
        
        total_stock += stock
        total_net += tot_net
        total_gross += tot_gross
        
    # Append totals row
    totals_row = [
        "ÖSSZESEN",
        "",
        "",
        "",
        total_stock,
        "",
        "",
        total_net,
        total_gross
    ]
    ws.append(totals_row)
    
    # Style totals row
    row_idx = len(products) + 2
    for col_num in range(1, len(totals_row) + 1):
        cell = ws.cell(row=row_idx, column=col_num)
        cell.font = Font(bold=True)
        if col_num in [5, 8, 9]:
            cell.alignment = Alignment(horizontal="right")
            
    # Set column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Freeze header
    ws.freeze_panes = "A2"
    
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=keszletertek.xlsx"}
    )


@router.get("/export/consistency")
async def export_consistency(
    db: AsyncSession = Depends(get_db)
):
    prod_stmt = select(Product)
    res = await db.execute(prod_stmt)
    products = res.scalars().all()
    
    mv_stmt = select(
        InventoryMovement.product_id,
        func.sum(InventoryMovement.quantity_delta).label("total_delta")
    ).group_by(InventoryMovement.product_id)
    mv_res = await db.execute(mv_stmt)
    movements_sum = {r[0]: r[1] for r in mv_res.all()}
    
    discrepancies = []
    for p in products:
        expected = movements_sum.get(p.id, 0)
        current = p.current_stock or 0
        
        if not p.is_archived and current != expected:
            discrepancies.append((p.name, p.sku or "", p.barcode, "Készlet eltérés", f"Terméktábla: {current} db, Mozgások: {expected} db", current, expected))
        if not p.is_archived and current < 0 and not p.allow_negative_stock:
            discrepancies.append((p.name, p.sku or "", p.barcode, "Tiltott negatív készlet", f"Készlet: {current} db", current, expected))
        if p.is_archived and current != 0:
            discrepancies.append((p.name, p.sku or "", p.barcode, "Archivált termék készlettel", f"Készlet: {current} db", current, expected))
            
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Konzisztencia Hiba Jelentés"
    
    headers = [
        "Terméknév",
        "SKU",
        "Vonalkód",
        "Hiba típusa",
        "Részletek",
        "Rendszer készlet",
        "Mozgások alapján várható"
    ]
    ws.append(headers)
    
    # Style headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="991B1B", end_color="991B1B", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")
        
    for row in discrepancies:
        ws.append(list(row))
        
    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    ws.freeze_panes = "A2"
    
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=keszlet_konzisztencia.xlsx"}
    )
