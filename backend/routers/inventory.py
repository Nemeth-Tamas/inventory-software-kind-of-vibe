from fastapi import APIRouter, Depends, HTTPException, status, File, UploadFile
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from database import get_db
from models import Product, User, UserRole
from models_inventory import InventoryMovement, MovementType
from auth import require_role, get_current_user
from audit_logger import log_audit
from event_bus import event_bus
from pydantic import BaseModel
from typing import Optional, List

router = APIRouter(prefix="/api/inventory", tags=["inventory"])


class ReceiptItem(BaseModel):
    product_id: str
    quantity: int
    purchase_price_net: Optional[int] = None
    purchase_price_gross: Optional[int] = None


class ReceiptRequest(BaseModel):
    items: List[ReceiptItem]
    supplier_id: Optional[str] = None
    location_id: str
    reference_number: Optional[str] = None
    note: Optional[str] = None
    idempotency_key: Optional[str] = None


class IssueItem(BaseModel):
    product_id: str
    quantity: int


class IssueRequest(BaseModel):
    items: List[IssueItem]
    location_id: str
    reason: str
    reference_number: Optional[str] = None
    note: Optional[str] = None
    idempotency_key: Optional[str] = None


class TransferRequest(BaseModel):
    product_id: str
    source_location_id: str
    destination_location_id: str
    quantity: int
    note: Optional[str] = None
    idempotency_key: Optional[str] = None


@router.post("/receipt", status_code=status.HTTP_201_CREATED)
async def goods_receipt(
    req: ReceiptRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        require_role([UserRole.ADMIN, UserRole.LEADER, UserRole.WAREHOUSE])
    ),
):
    try:
        # Check idempotency
        if req.idempotency_key:
            dup_res = await db.execute(
                select(InventoryMovement).where(
                    InventoryMovement.idempotency_key == req.idempotency_key
                )
            )
            if dup_res.scalar_one_or_none():
                raise HTTPException(
                    status_code=400,
                    detail="Ez a bevételezés már fel lett dolgozva (duplikált idempotens kérés)!",
                )

        for item in req.items:
            result = await db.execute(
                select(Product).where(Product.id == item.product_id).with_for_update()
            )
            product = result.scalar_one_or_none()
            if not product:
                raise HTTPException(
                    status_code=404, detail=f"Termék nem található: {item.product_id}"
                )

            stock_before = product.current_stock
            product.current_stock += item.quantity

            if item.purchase_price_net is not None:
                product.purchase_price_net = item.purchase_price_net
            if item.purchase_price_gross is not None:
                product.purchase_price_gross = item.purchase_price_gross

            movement = InventoryMovement(
                product_id=product.id,
                quantity_delta=item.quantity,
                stock_before=stock_before,
                stock_after=product.current_stock,
                destination_location_id=req.location_id,
                supplier_id=req.supplier_id,
                price_net=item.purchase_price_net,
                movement_type=MovementType.RECEIPT,
                reason="Bevételezés",
                reference_number=req.reference_number,
                user_id=current_user.id,
                note=req.note,
                idempotency_key=req.idempotency_key,
            )
            db.add(movement)
            await log_audit(
                db,
                current_user.id,
                current_user.username,
                f"Készlet bevételezés: {product.name} (+{item.quantity} db)",
            )

        await db.commit()
    except Exception as e:
        await db.rollback()
        raise e

    await event_bus.publish(
        "inventory_events", "stock_updated", {"message": "Készlet bevételezés történt"}
    )
    return {"status": "success", "message": "Bevételezés sikeresen rögzítve"}


@router.post("/issue", status_code=status.HTTP_201_CREATED)
async def stock_issue(
    req: IssueRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        require_role(
            [UserRole.ADMIN, UserRole.LEADER, UserRole.WAREHOUSE, UserRole.SALES]
        )
    ),
):
    try:
        # Check idempotency
        if req.idempotency_key:
            dup_res = await db.execute(
                select(InventoryMovement).where(
                    InventoryMovement.idempotency_key == req.idempotency_key
                )
            )
            if dup_res.scalar_one_or_none():
                raise HTTPException(
                    status_code=400,
                    detail="Ez a kiadás már fel lett dolgozva (duplikált idempotens kérés)!",
                )

        for item in req.items:
            result = await db.execute(
                select(Product).where(Product.id == item.product_id).with_for_update()
            )
            product = result.scalar_one_or_none()
            if not product:
                raise HTTPException(
                    status_code=404, detail=f"Termék nem található: {item.product_id}"
                )

            if (
                not product.allow_negative_stock
                and product.current_stock < item.quantity
            ):
                raise HTTPException(
                    status_code=400,
                    detail=f"Nincs elegendő készlet a termékből: {product.name} (Elérhető: {product.current_stock})",
                )

            stock_before = product.current_stock
            product.current_stock -= item.quantity

            movement = InventoryMovement(
                product_id=product.id,
                quantity_delta=-item.quantity,
                stock_before=stock_before,
                stock_after=product.current_stock,
                source_location_id=req.location_id,
                movement_type=MovementType.SALE,
                reason=req.reason,
                reference_number=req.reference_number,
                user_id=current_user.id,
                note=req.note,
                idempotency_key=req.idempotency_key,
            )
            db.add(movement)
            await log_audit(
                db,
                current_user.id,
                current_user.username,
                f"Készlet kiadás: {product.name} (-{item.quantity} db, indok: {req.reason})",
            )

        await db.commit()
    except Exception as e:
        await db.rollback()
        raise e

    await event_bus.publish(
        "inventory_events", "stock_updated", {"message": "Készlet kiadás történt"}
    )
    return {"status": "success", "message": "Kiadás sikeresen rögzítve"}


@router.post("/transfer", status_code=status.HTTP_201_CREATED)
async def stock_transfer(
    req: TransferRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        require_role([UserRole.ADMIN, UserRole.LEADER, UserRole.WAREHOUSE])
    ),
):
    try:
        # Check idempotency
        if req.idempotency_key:
            dup_res = await db.execute(
                select(InventoryMovement).where(
                    InventoryMovement.idempotency_key == req.idempotency_key
                )
            )
            if dup_res.scalar_one_or_none():
                raise HTTPException(
                    status_code=400,
                    detail="Ez az átadás már fel lett dolgozva (duplikált idempotens kérés)!",
                )

        result = await db.execute(
            select(Product).where(Product.id == req.product_id).with_for_update()
        )
        product = result.scalar_one_or_none()
        if not product:
            raise HTTPException(status_code=404, detail="Termék nem található")

        if not product.allow_negative_stock and product.current_stock < req.quantity:
            raise HTTPException(
                status_code=400,
                detail=f"Nincs elegendő készlet az átadáshoz: {product.name} (Készleten: {product.current_stock})",
            )

        stock_before = product.current_stock
        movement = InventoryMovement(
            product_id=product.id,
            quantity_delta=0,
            stock_before=stock_before,
            stock_after=stock_before,
            source_location_id=req.source_location_id,
            destination_location_id=req.destination_location_id,
            movement_type=MovementType.TRANSFER,
            reason="Helyszínek közötti átadás",
            user_id=current_user.id,
            note=req.note,
            idempotency_key=req.idempotency_key,
        )
        db.add(movement)
        await log_audit(
            db,
            current_user.id,
            current_user.username,
            f"Készlet átadás: {product.name} ({req.quantity} db átadva)",
        )
        await db.commit()
    except Exception as e:
        await db.rollback()
        raise e

    await event_bus.publish(
        "inventory_events", "stock_updated", {"message": "Készletátadás történt"}
    )
    return {"status": "success", "message": "Átadás sikeresen rögzítve"}


@router.get("/movements")
async def get_movements(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(InventoryMovement)
        .options(
            selectinload(InventoryMovement.product),
            selectinload(InventoryMovement.user),
            selectinload(InventoryMovement.supplier),
        )
        .order_by(InventoryMovement.timestamp.desc())
        .limit(100)
    )
    movements = result.scalars().all()
    return [
        {
            "id": m.id,
            "product_name": m.product.name if m.product else "Ismeretlen termék",
            "barcode": m.product.barcode if m.product else "",
            "quantity_delta": m.quantity_delta,
            "stock_before": m.stock_before,
            "stock_after": m.stock_after,
            "movement_type": m.movement_type.value,
            "reason": m.reason,
            "timestamp": m.timestamp.isoformat(),
            "user": m.user.username if m.user else "Rendszer",
            "supplier_name": m.supplier.name if m.supplier else "",
            "price_net": m.price_net,
        }
        for m in movements
    ]


class CheckMovementsRequest(BaseModel):
    product_ids: List[str]


class OpeningStockItem(BaseModel):
    product_id: str
    quantity: int
    location_id: str


class OpeningStockRequest(BaseModel):
    items: List[OpeningStockItem]
    idempotency_key: Optional[str] = None
    note: Optional[str] = None
    force_apply: Optional[bool] = False


@router.post("/opening-stock/check-movements")
async def check_movements(
    req: CheckMovementsRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        require_role([UserRole.ADMIN, UserRole.LEADER, UserRole.WAREHOUSE])
    ),
):
    from sqlalchemy import func

    result = await db.execute(
        select(InventoryMovement.product_id, func.count(InventoryMovement.id))
        .where(InventoryMovement.product_id.in_(req.product_ids))
        .group_by(InventoryMovement.product_id)
    )
    counts = {row[0]: row[1] for row in result.all()}
    return {pid: counts.get(pid, 0) > 0 for pid in req.product_ids}


@router.post("/opening-stock")
async def apply_opening_stock(
    req: OpeningStockRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        require_role([UserRole.ADMIN, UserRole.LEADER, UserRole.WAREHOUSE])
    ),
):
    from models import Location

    try:
        product_ids = [item.product_id for item in req.items]
        location_ids = [item.location_id for item in req.items]

        products_res = await db.execute(
            select(Product).where(Product.id.in_(product_ids))
        )
        products_map = {p.id: p for p in products_res.scalars().all()}

        locations_res = await db.execute(
            select(Location).where(Location.id.in_(location_ids))
        )
        locations_map = {loc.id: loc for loc in locations_res.scalars().all()}

        # Check if any product has existing movements
        if not req.force_apply:
            m_res = await db.execute(
                select(InventoryMovement.product_id)
                .where(InventoryMovement.product_id.in_(product_ids))
                .group_by(InventoryMovement.product_id)
            )
            products_with_movements = [row[0] for row in m_res.all()]
            if products_with_movements:
                names = [
                    products_map[pid].name
                    for pid in products_with_movements
                    if pid in products_map
                ]
                raise HTTPException(
                    status_code=409,
                    detail={
                        "error_type": "EXISTING_MOVEMENTS",
                        "message": "Egyes termékek már rendelkeznek készletmozgással. Vezetői jóváhagyás szükséges.",
                        "products": names,
                    },
                )

        if req.force_apply and current_user.role not in [
            UserRole.ADMIN,
            UserRole.LEADER,
        ]:
            # Double check if there actually are existing movements first
            m_res = await db.execute(
                select(InventoryMovement.product_id)
                .where(InventoryMovement.product_id.in_(product_ids))
                .group_by(InventoryMovement.product_id)
            )
            products_with_movements = [row[0] for row in m_res.all()]
            if products_with_movements:
                raise HTTPException(
                    status_code=403,
                    detail="Csak adminisztrátor vagy vezető hagyhatja jóvá a már készletmozgással rendelkező termékek nyitókészletének módosítását.",
                )

        # Apply movements
        for item in req.items:
            product = products_map.get(item.product_id)
            if not product:
                raise HTTPException(
                    status_code=400, detail=f"Termék nem található: {item.product_id}"
                )

            location = locations_map.get(item.location_id)
            if not location:
                raise HTTPException(
                    status_code=400, detail=f"Tárhely nem található: {item.location_id}"
                )

            stock_before = product.current_stock or 0
            stock_after = item.quantity
            delta = stock_after - stock_before

            mv = InventoryMovement(
                product_id=product.id,
                quantity_delta=delta,
                stock_before=stock_before,
                stock_after=stock_after,
                destination_location_id=location.id,
                movement_type=MovementType.OPENING,
                note=req.note or "Nyitókészlet rögzítése",
                user_id=current_user.id,
            )

            product.current_stock = stock_after
            db.add(mv)

        await log_audit(
            db,
            current_user.id,
            current_user.username,
            f"Nyitókészlet rögzítve: {len(req.items)} db termékhez.",
        )
        await db.commit()
    except Exception as e:
        await db.rollback()
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(
            status_code=500, detail=f"Hiba a nyitókészlet rögzítése során: {str(e)}"
        )

    await event_bus.publish(
        "inventory_events", "stock_updated", {"message": "Nyitókészlet rögzítve"}
    )
    return {
        "status": "success",
        "message": f"Nyitókészlet sikeresen rögzítve {len(req.items)} termékhez.",
    }


@router.get("/opening-stock/template")
async def download_opening_stock_template():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Nyitókészlet Sablon"

    # Headers
    headers = [
        "Belső vonalkód *",
        "Cikkszám (SKU)",
        "Terméknév",
        "Nyitó mennyiség *",
        "Tárhely (Név) *",
    ]

    # Styling
    header_fill = PatternFill(
        start_color="0284C7", end_color="0284C7", fill_type="solid"
    )
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    thin_side = Side(border_style="thin", color="CBD5E1")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Example rows
    examples = [
        ["001234", "SKU-MON-27", "Test Monitor 27", 15, "Fő raktár"],
        ["005678", "SKU-TAR-1TB", "1 TB Felhő alapú tárhely", 5, "Iroda"],
    ]

    for row_idx, example in enumerate(examples, 2):
        for col_idx, val in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="left" if col_idx != 4 else "right")
            cell.border = border
            if col_idx == 1:
                cell.number_format = "@"

    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 15)

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=nyitokeszlet_sablon.xlsx"
        },
    )


@router.post("/opening-stock/import-preview")
async def import_preview(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        require_role([UserRole.ADMIN, UserRole.LEADER, UserRole.WAREHOUSE])
    ),
):
    import io
    from openpyxl import load_workbook
    from models import Location

    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=400, detail="Az Excel fájl üres.")

        header = rows[0]
        barcode_idx = -1
        sku_idx = -1
        name_idx = -1
        qty_idx = -1
        loc_idx = -1

        for idx, h in enumerate(header):
            if h is None:
                continue
            h_clean = str(h).lower().strip()
            if "vonalkód" in h_clean or "barcode" in h_clean:
                barcode_idx = idx
            elif "sku" in h_clean or "cikkszám" in h_clean:
                sku_idx = idx
            elif "terméknév" in h_clean or "megnevezés" in h_clean or "név" in h_clean:
                name_idx = idx
            elif "mennyiség" in h_clean or "darab" in h_clean or "qty" in h_clean:
                qty_idx = idx
            elif "tárhely" in h_clean or "location" in h_clean:
                loc_idx = idx

        if barcode_idx == -1 or qty_idx == -1 or loc_idx == -1:
            raise HTTPException(
                status_code=400,
                detail="Nem találhatók a kötelező oszlopok: 'Belső vonalkód', 'Nyitó mennyiség' és 'Tárhely'.",
            )

        prod_res = await db.execute(select(Product))
        all_products = prod_res.scalars().all()
        products_by_barcode = {p.barcode: p for p in all_products if p.barcode}

        loc_res = await db.execute(select(Location))
        all_locations = loc_res.scalars().all()
        locations_by_name = {loc.name.lower().strip(): loc for loc in all_locations}
        locations_by_id = {loc.id: loc for loc in all_locations}

        preview_rows = []
        seen_barcodes = set()
        has_duplicates = False

        for row_idx, r in enumerate(rows[1:], 2):
            if r is None or all(cell is None for cell in r):
                continue

            raw_barcode = (
                str(r[barcode_idx]).strip()
                if (barcode_idx < len(r) and r[barcode_idx] is not None)
                else ""
            )
            if raw_barcode.endswith(".0"):  # Fix floats read as string
                raw_barcode = raw_barcode[:-2]
            if raw_barcode.isdigit() and len(raw_barcode) < 6:
                raw_barcode = raw_barcode.zfill(6)

            raw_sku = (
                str(r[sku_idx]).strip()
                if (sku_idx < len(r) and r[sku_idx] is not None)
                else ""
            )
            raw_name = (
                str(r[name_idx]).strip()
                if (name_idx < len(r) and r[name_idx] is not None)
                else ""
            )

            raw_qty_val = r[qty_idx]
            raw_loc = (
                str(r[loc_idx]).strip()
                if (loc_idx < len(r) and r[loc_idx] is not None)
                else ""
            )

            errors = []
            product = None
            location = None

            if not raw_barcode:
                errors.append("A belső vonalkód hiányzik.")
            else:
                product = products_by_barcode.get(raw_barcode)
                if not product:
                    errors.append(f"Ismeretlen termék (vonalkód: {raw_barcode}).")

            if raw_qty_val is None:
                errors.append("A mennyiség hiányzik.")
                qty = 0
            else:
                try:
                    # Handle float string from Excel representation of integers
                    if isinstance(raw_qty_val, float):
                        qty = int(raw_qty_val)
                    elif isinstance(raw_qty_val, str) and "." in raw_qty_val:
                        qty = int(float(raw_qty_val))
                    else:
                        qty = int(raw_qty_val)
                    if qty < 0:
                        errors.append("A mennyiség nem lehet negatív.")
                except ValueError:
                    errors.append("A mennyiség érvénytelen egész szám.")
                    qty = 0

            if not raw_loc:
                errors.append("A tárhely hiányzik.")
            else:
                location = locations_by_name.get(raw_loc.lower())
                if not location:
                    location = locations_by_id.get(raw_loc)
                if not location:
                    errors.append(f"Ismeretlen tárhely: {raw_loc}.")

            is_duplicate = False
            if raw_barcode and raw_barcode in seen_barcodes:
                is_duplicate = True
                has_duplicates = True
                errors.append("Ez a vonalkód többször szerepel a fájlban.")
            elif raw_barcode:
                seen_barcodes.add(raw_barcode)

            preview_rows.append(
                {
                    "row_index": row_idx,
                    "barcode": raw_barcode,
                    "sku": raw_sku or (product.sku if product else ""),
                    "name": raw_name or (product.name if product else ""),
                    "quantity": qty,
                    "location_name": raw_loc,
                    "location_id": location.id if location else None,
                    "product_id": product.id if product else None,
                    "errors": errors,
                    "is_valid": len(errors) == 0,
                    "is_duplicate": is_duplicate,
                }
            )

        return {
            "has_errors": any(not r["is_valid"] for r in preview_rows),
            "has_duplicates": has_duplicates,
            "items": preview_rows,
        }
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(
            status_code=500, detail=f"Hiba a fájl feldolgozása során: {str(e)}"
        )


class ValuationItem(BaseModel):
    product_id: str
    product_name: str
    sku: Optional[str] = None
    category_name: Optional[str] = None
    current_stock: int
    purchase_price_net: int
    purchase_price_gross: int
    total_value_net: int
    total_value_gross: int
    location_name: Optional[str]
    price_warning: bool


class ValuationResponse(BaseModel):
    items: List[ValuationItem]
    total_stock: int
    total_value_net: int
    total_value_gross: int


@router.get("/valuation", response_model=ValuationResponse)
async def get_valuation(
    category_id: Optional[str] = None,
    location_id: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    stmt = select(Product).where(Product.is_archived.is_(False))
    if category_id:
        stmt = stmt.where(Product.category_id == category_id)
    if location_id:
        stmt = stmt.where(Product.default_location_id == location_id)

    stmt = stmt.options(
        selectinload(Product.category), selectinload(Product.default_location)
    )

    result = await db.execute(stmt)
    products = result.scalars().all()

    items = []
    total_stock = 0
    total_value_net = 0
    total_value_gross = 0

    for p in products:
        stock = p.current_stock or 0
        price_net = p.purchase_price_net or 0
        price_gross = p.purchase_price_gross or 0

        # If one of the prices is 0, estimate it
        if price_net > 0 and price_gross == 0:
            price_gross = int(price_net * 1.27)
        elif price_gross > 0 and price_net == 0:
            price_net = int(price_gross / 1.27)

        tot_net = stock * price_net
        tot_gross = stock * price_gross

        price_warning = (price_net <= 0) or (price_gross <= 0)

        items.append(
            ValuationItem(
                product_id=p.id,
                product_name=p.name,
                sku=p.sku,
                category_name=p.category.name if p.category else "Nincs kategória",
                current_stock=stock,
                purchase_price_net=price_net,
                purchase_price_gross=price_gross,
                total_value_net=tot_net,
                total_value_gross=tot_gross,
                location_name=p.default_location.name
                if p.default_location
                else "Nincs helyszín",
                price_warning=price_warning,
            )
        )

        total_stock += stock
        total_value_net += tot_net
        total_value_gross += tot_gross

    return ValuationResponse(
        items=items,
        total_stock=total_stock,
        total_value_net=total_value_net,
        total_value_gross=total_value_gross,
    )


class ConsistencyDiscrepancy(BaseModel):
    product_id: str
    product_name: str
    barcode: str
    sku: Optional[str] = None
    type: str  # "stock_mismatch", "negative_stock", "archived_with_stock"
    details: str
    current_stock: int
    expected_stock: int


class ConsistencyReport(BaseModel):
    has_issues: bool
    discrepancies: List[ConsistencyDiscrepancy]


@router.get("/consistency", response_model=ConsistencyReport)
async def get_stock_consistency(
    db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)
):
    prod_stmt = select(Product)
    res = await db.execute(prod_stmt)
    products = res.scalars().all()

    mv_stmt = select(
        InventoryMovement.product_id,
        func.sum(InventoryMovement.quantity_delta).label("total_delta"),
    ).group_by(InventoryMovement.product_id)
    mv_res = await db.execute(mv_stmt)
    movements_sum = {r[0]: r[1] for r in mv_res.all()}

    discrepancies = []

    for p in products:
        expected = movements_sum.get(p.id, 0)
        current = p.current_stock or 0

        # Check 1: Stock mismatch
        if not p.is_archived and current != expected:
            discrepancies.append(
                ConsistencyDiscrepancy(
                    product_id=p.id,
                    product_name=p.name,
                    barcode=p.barcode,
                    sku=p.sku,
                    type="stock_mismatch",
                    details=f"Készlet eltérés: a terméktáblában {current} db szerepel, de a mozgások összege {expected} db.",
                    current_stock=current,
                    expected_stock=expected,
                )
            )

        # Check 2: Negative stock where not allowed
        if not p.is_archived and current < 0 and not p.allow_negative_stock:
            discrepancies.append(
                ConsistencyDiscrepancy(
                    product_id=p.id,
                    product_name=p.name,
                    barcode=p.barcode,
                    sku=p.sku,
                    type="negative_stock",
                    details=f"Negatív készlet: a jelenlegi készlet {current} db, de a negatív készlet nincs engedélyezve a terméknél.",
                    current_stock=current,
                    expected_stock=expected,
                )
            )

        # Check 3: Archived product with non-zero stock
        if p.is_archived and current != 0:
            discrepancies.append(
                ConsistencyDiscrepancy(
                    product_id=p.id,
                    product_name=p.name,
                    barcode=p.barcode,
                    sku=p.sku,
                    type="archived_with_stock",
                    details=f"Archivált termék aktív készlettel: a termék archiválva van, de a készlete {current} db.",
                    current_stock=current,
                    expected_stock=expected,
                )
            )

    return ConsistencyReport(
        has_issues=len(discrepancies) > 0, discrepancies=discrepancies
    )
